from typing import Dict, Any

from openpyxl import load_workbook
from openpyxl.chart import Reference, BarChart

workbook = load_workbook('mediumAircraftData.xlsx')
sheet = workbook['Sheet1']
AnimalDict = {}
yearDict = {}
newyearDict = {}
monthDict = {}
airlineDict = {}
newairlineDict = {}
# this for loop replaces the names of the animals from specific to general names
def Species():
    global AnimalDict
    for rowNum in range(2, sheet.max_row):
        animal = sheet.cell(row=rowNum + 1, column=32).value
        if animal == 'UNKNOWN MEDIUM BIRD':
            sheet.cell(row=rowNum + 1, column=32, value='BIRD')
        elif animal == 'BARN OWL':
            sheet.cell(row=rowNum + 1, column=32, value='OWL')
        elif animal == 'PACIFIC GOLDEN-PLOVER':
            sheet.cell(row=rowNum + 1, column=32, value='PLOVER')
        elif animal == 'MUNIAS':
            sheet.cell(row=rowNum + 1, column=32, value='MUNIAS')
        elif animal == 'HORNED LARK':
            sheet.cell(row=rowNum + 1, column=32, value='LARK')
        elif animal == 'HORNED OWL':
            sheet.cell(row=rowNum + 1, column=32, value='OWL')
        elif animal == 'ROCK PIGEON':
            sheet.cell(row=rowNum + 1, column=32, value='PIGEON')
        elif animal == 'SPOTTED DOVE':
            sheet.cell(row=rowNum + 1, column=32, value='DOVE')
        elif animal == 'HOUSE SPARROW':
            sheet.cell(row=rowNum + 1, column=32, value='SPARROW')
        elif animal == 'BARn OWL':
            sheet.cell(row=rowNum + 1, column=32, value='OWL')
        elif animal == 'WHITE-TAILED DEER':
            sheet.cell(row=rowNum + 1, column=32, value='DEER')
        elif animal == 'UNKNOWN BIRD':
            sheet.cell(row=rowNum + 1, column=32, value='UNK BIRD')
        elif animal == 'UNKNOWN SMALL BIRD':
            sheet.cell(row=rowNum + 1, column=32, value='SMALL BIRD')
        elif animal == 'EUROPEAN STARLING':
            sheet.cell(row=rowNum + 1, column=32, value='STARLING')
        elif animal == 'MOURNING DOVE':
            sheet.cell(row=rowNum + 1, column=32, value='DOVE')
        elif animal == 'KILLDEER':
            sheet.cell(row=rowNum + 1, column=32, value='DEER')
        elif animal == 'AMERICAN KESTREL':
            sheet.cell(row=rowNum + 1, column=32, value='KESTREL')
        elif animal == 'BARN SWALLOW':
            sheet.cell(row=rowNum + 1, column=32, value='SWALLOW')

    # this code goes through the column 32 and counts the number of each animal incidence
    for row in sheet.iter_rows(min_row=2, min_col=32, max_col=32, values_only=True):
        if row[0] in AnimalDict:
            AnimalDict[row[0]] += 1
        else:
            AnimalDict[row[0]] = 1


Species()  # this is so that the above defined function is called upon


def largevalues(input):
    global newDict
    newDict = {}
    maxvalue = max(input, key=input.get)
    percent = input[maxvalue] * 0.10
    for key in input:
        if input[key] >= percent:
            newDict[key] = input[key]


largevalues(
    AnimalDict)  # this code filters through the dictionary and only adds the values that are 10% larger than the
# highest value.

newSheet = workbook.create_sheet('chartForAnimals')

i = 2
for key in newDict:
    newSheet.cell(row=i, column=1).value = key
    newSheet.cell(row=i, column=2).value = newDict[key]
    i += 1
newSheet.cell(row=1, column=1, value='Animal')
newSheet.cell(row=1, column=2, value='Frequency')

chart = BarChart()
values = Reference(worksheet=newSheet,
                   min_row=1,
                   max_row=newSheet.max_row,
                   min_col=1,
                   max_col=2)
animalLabel = Reference(worksheet=newSheet, min_row=2, max_row=newSheet.max_row, min_col=1)
chart.add_data(values, titles_from_data=True)
chart.set_categories(animalLabel)
chart.title = 'SPECIES'
newSheet.add_chart(chart, "E2")  # this takes information from newSheet and uses that data to create a Bar chart
print(newDict)


def years():
    global yearDict
    for rowNum in range(2, sheet.max_row):
        year = sheet.cell(row=rowNum + 1, column=2).value

    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
        if row[0] in yearDict:
            yearDict[row[0]] += 1
        else:
            yearDict[row[0]] = 1


years()


# this code goes through the year column and counts the number of years and adds it to the dictionary

def bigyears(input):
    global newyearDict
    maxvalue = max(input, key=input.get)
    percent = input[maxvalue] * 0.10
    for key in input:
        if input[key] >= percent:
            newyearDict[key] = input[key]


bigyears(yearDict)

yearssheet = workbook.create_sheet('chartForYears')

i = 2
for key in newyearDict:
    yearssheet.cell(row=i, column=1).value = key
    yearssheet.cell(row=i, column=2).value = newyearDict[key]
    i += 1
yearssheet.cell(row=1, column=1, value='Year')
yearssheet.cell(row=1, column=2, value='Occurrence')
chart = BarChart()
values = Reference(worksheet=yearssheet,
                   min_row=1,
                   max_row=yearssheet.max_row,
                   min_col=2,
                   max_col=2)
yearLabel = Reference(worksheet=yearssheet, min_row=2, max_row=yearssheet.max_row, min_col=1)
chart.add_data(values, titles_from_data=True)
chart.set_categories(yearLabel)
chart.title = 'YEARS'
yearssheet.add_chart(chart, "E2")
print(newyearDict)

# takes the information from yearssheet and makes a barchart from it




def month():
    global monthDict
    for rowNum in range(2, sheet.max_row):
        month = sheet.cell(row=rowNum + 1, column=3).value
    for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True):
        if row[0] in monthDict:
            monthDict[row[0]] += 1
        else:
            monthDict[row[0]] = 1

month()

monthsheet = workbook.create_sheet('chartForMonths')
i = 2
for key in monthDict:
    monthsheet.cell(row=i, column=1).value = key
    monthsheet.cell(row=i, column=2).value = monthDict[key]
    i += 1
monthsheet.cell(row=1, column=1, value='Month')
monthsheet.cell(row=1, column=2, value='Occurrence')

chart = BarChart()
values = Reference(worksheet=monthsheet,
                   min_row=1,
                   max_row=monthsheet.max_row,
                   min_col=2,
                   max_col=2)
monthLabel = Reference(worksheet=monthsheet, min_row=2, max_row=monthsheet.max_row, min_col=1)
chart.add_data(values, titles_from_data=True)
chart.set_categories(monthLabel)
chart.title = 'MONTH'
monthsheet.add_chart(chart, "E2")
print(monthDict)
# takes the information from the monthsheet and turns it into a bar chart

def airlines():
    global airlineDict
    for rowNum in range(2, sheet.max_row):
        airline = sheet.cell(row=rowNum + 1, column=6).value

    for row in sheet.iter_rows(min_row=2, min_col=6, max_col=6, values_only=True):
        if row[0] in airlineDict:
           airlineDict[row[0]] += 1
        else:
            airlineDict[row[0]] = 1


airlines()


# this code goes through the year column and counts the number of airlines and adds it to the dictionary

def bigairlines(input):
    global newairlineDict
    maxvalue = max(input, key=input.get)
    percent = input[maxvalue] * 0.10
    for key in input:
        if input[key] >= percent:
            newairlineDict[key] = input[key]


bigairlines(airlineDict)

airlinesheet = workbook.create_sheet('chartForAirlines')

i = 2
myKeys = list(newairlineDict.keys())
myKeys.sort()
sorted_dict = {i: newairlineDict[i] for i in myKeys}
for key in sorted_dict:
    airlinesheet.cell(row=i, column=1).value = key
    airlinesheet.cell(row=i, column=2).value = sorted_dict[key]
    i += 1
airlinesheet.cell(row=1, column=1, value='airline')
airlinesheet.cell(row=1, column=2, value='Occurrence')

print(sorted_dict)
chart = BarChart()
values = Reference(worksheet=airlinesheet,
                   min_row=1,
                   max_row=airlinesheet.max_row,
                   min_col=2,
                   max_col=2)
airlineLabel = Reference(worksheet=airlinesheet, min_row=2, max_row=airlinesheet.max_row, min_col=1)
chart.add_data(values, titles_from_data=True)
chart.set_categories(airlineLabel)
chart.title = 'AIRLINES'
airlinesheet.add_chart(chart, "E2")
workbook.save('mediumAircraftData.xlsx')
